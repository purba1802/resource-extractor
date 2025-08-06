import streamlit as st
import cv2
import numpy as np
from openpyxl import load_workbook
import os

EXCEL_FILE = "Output.xlsx"

# Fungsi sederhana: convert gambar ke angka via segmentasi digit (dummy extractor)
def extract_digits(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    digits = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        if 10 < h < 80:  # filter objek kecil / besar
            roi = thresh[y:y+h, x:x+w]
            digits.append((x, roi))

    digits = sorted(digits, key=lambda x: x[0])  # urut dari kiri
    return len(digits)  # contoh sederhana: hitung jumlah blok

# --- Inisialisasi Session State ---
if "kingdom" not in st.session_state:
    st.session_state["kingdom"] = ""
if "excel_initialized" not in st.session_state:
    if not os.path.exists(EXCEL_FILE):
        from shutil import copyfile
        copyfile("Template.xlsx", EXCEL_FILE)
    st.session_state["excel_initialized"] = True

st.title("Resource Extractor (Tanpa OCR Eksternal)")

# --- Input Kingdom (hanya sekali) ---
if st.session_state["kingdom"] == "":
    kingdom = st.text_input("Masukkan Kingdom:")
    if st.button("Konfirmasi Kingdom"):
        st.session_state["kingdom"] = kingdom
else:
    st.success(f"Kingdom: {st.session_state['kingdom']} (disimpan di K3)")

    # --- Input Data Akun ---
    nick = st.text_input("Nick:")
    tp = st.text_input("Level Trading Post:")
    sh = st.text_input("Level Store House:")
    uploaded_file = st.file_uploader("Upload screenshot total resource", type=["png", "jpg", "jpeg"])

    # --- Slider Crop Area ---
    st.subheader("Atur Area Crop untuk Deteksi Angka")
    top_crop = st.slider("Atas (%)", 0, 100, 40)
    bottom_crop = st.slider("Bawah (%)", 0, 100, 100)
    left_crop = st.slider("Kiri (%)", 0, 100, 50)
    right_crop = st.slider("Kanan (%)", 0, 100, 100)

    # --- Tombol Tambah Data ---
    if st.button("Tambah Data"):
        if uploaded_file is not None and nick != "":
            # Baca gambar
            file_bytes = np.asarray(bytearray(uploaded_file.read()), dtype=np.uint8)
            img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)

            # Crop gambar
            h, w, _ = img.shape
            crop_img = img[int(h*top_crop/100):int(h*bottom_crop/100),
                           int(w*left_crop/100):int(w*right_crop/100)]

            # Gunakan fungsi sederhana (hanya mendeteksi jumlah blok angka)
            value = extract_digits(crop_img)

            resources = {"Food": value*1000, "Wood": value*2000, "Stone": value*3000, "Gold": value*4000}

            # Update Excel
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws["K3"] = st.session_state["kingdom"]

            # Cari baris kosong mulai A11
            row = 11
            while ws.cell(row=row, column=3).value not in [None, ""]:
                row += 1

            ws.cell(row=row, column=1, value=row-10)  # A: nomor
            ws.cell(row=row, column=2, value="")      # B: KODE
            ws.cell(row=row, column=3, value=nick)    # C: Nick
            ws.cell(row=row, column=4, value=tp)      # D: Trading Post
            ws.cell(row=row, column=6, value=sh)      # F: Store House
            ws.cell(row=row, column=7, value=resources["Food"])   # G: Food
            ws.cell(row=row, column=8, value=resources["Wood"])   # H: Wood
            ws.cell(row=row, column=9, value=resources["Stone"])  # I: Stone
            ws.cell(row=row, column=10, value=resources["Gold"])  # J: Gold

            wb.save(EXCEL_FILE)
            st.success(f"Data {nick} berhasil ditambahkan.")

            # Reset form
            st.rerun()
        else:
            st.error("Nick dan screenshot harus diisi.")

    # --- Tombol Download Excel ---
    if st.button("Download Excel"):
        with open(EXCEL_FILE, "rb") as f:
            st.download_button(label="Klik untuk mengunduh Output.xlsx",
                               data=f,
                               file_name="Output.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
